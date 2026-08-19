from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import (
    Actividad,
    ActividadArea,
    ActividadCriterio,
    ActividadCriterioCompetencia,
    ActividadMaquina,
    Area,
    Cargo,
    Competencia,
    CriterioIndicador,
    ImportacionMatriz,
    Maquina,
    MaquinaProceso,
    MatrizPuestoVersion,
    Puesto,
    PuestoActividad,
    PuestoActividadCompetencia,
    PuestoActividadCriterio,
    PuestoActividadCriterioRequisito,
    PuestoMaquina,
)
from app.services.codigos import generar_codigo

AREA_NAME = "Operación Línea Principal"
LEGACY_POSITION_NAME = "Operador Línea Principal"
IMPORTER_VERSION = "3"

ROLE_CONFIG = {
    "operador": {"cargo": "Operador", "prefix": "Operador", "default_level": 3},
    "ayudante": {"cargo": "Ayudante", "prefix": "Ayudante", "default_level": 2},
}

MACRO_COMPETENCIES = {
    "operacion": ("Operación y proceso", "tecnica", False),
    "inspeccion": ("Inspección y diagnóstico", "tecnica", False),
    "calidad": ("Calidad", "calidad", False),
    "seguridad": ("Seguridad", "seguridad", True),
    "coordinacion": ("Coordinación", "coordinacion", False),
    "conductuales": ("Conductuales", "conductual", False),
}


def is_operator_profile(content: bytes) -> bool:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return False
    try:
        return _find_profile_sheets(workbook) is not None
    finally:
        workbook.close()


def _text(value: Any) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _split(value: Any) -> list[str]:
    return [
        item.strip().rstrip(".") for item in str(value or "").split(";") if item.strip()
    ]


def _number(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}".rstrip("0").rstrip(".")
    return str(value or "").strip()


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")


def _find_profile_sheets(workbook) -> dict[str, str] | None:
    activities = None
    summary = None
    criteria = None
    for name in workbook.sheetnames:
        normalized = _slug(name)
        if normalized in {"perfil-operador-principal", "actividades-del-cargo"}:
            activities = name
        elif normalized in {"resumen-competencias", "resumen-de-competencias"}:
            summary = name
        elif normalized == "criterios-individualizados":
            criteria = name
    if activities and summary and criteria:
        return {"activities": activities, "summary": summary, "criteria": criteria}
    return None


def _profile_namespace(source: str, title: str) -> tuple[str, str]:
    match = re.search(r"MP-PO-TS\d+-ASE-\d+", source, re.IGNORECASE)
    procedure = match.group(0).upper() if match else _slug(title).upper()
    return procedure, procedure.replace("-", "")


def _macro_keys(row: tuple[Any, ...]) -> set[str]:
    activity = str(row[2] or "")
    criterion = str(row[3] or "")
    technical = str(row[4] or "")
    behavioral = str(row[5] or "")
    complete = " ".join((activity, criterion, technical, behavioral)).casefold()
    keys: set[str] = set()

    if technical:
        inspection_tokens = (
            "inspecc",
            "verific",
            "alarma",
            "diagnóst",
            "diagnost",
            "fotocelda",
            "fuga",
        )
        keys.add(
            "inspeccion"
            if any(token in complete for token in inspection_tokens)
            else "operacion"
        )
    if any(
        token in complete
        for token in (
            "medida",
            "dimensional",
            "calidad",
            "patrón de corte",
            "patron de corte",
        )
    ):
        keys.add("calidad")
    if behavioral:
        keys.add("conductuales")
    if any(
        token in complete
        for token in (
            "coordina",
            "comunica",
            "informa",
            "radio",
            "supervisor",
            "ayudante",
            "mantención",
            "mantencion",
        )
    ):
        keys.add("coordinacion")
    return keys


def parse_operator_profile(content: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    try:
        sheets = _find_profile_sheets(workbook)
        if sheets is None:
            raise ValueError(
                "El archivo no contiene hojas de actividades, resumen y criterios "
                "con una estructura reconocida"
            )
        activity_sheet = workbook[sheets["activities"]]
        criterion_sheet = workbook[sheets["criteria"]]
        title = _text(activity_sheet["A1"].value) or "Perfil operativo"
        source = _text(activity_sheet["A2"].value) or "Fuente no informada"
        procedure, namespace = _profile_namespace(source, title)

        activities: dict[str, dict[str, Any]] = {}
        for order, row in enumerate(list(activity_sheet.values)[4:]):
            if len(row) < 8 or not row[2]:
                continue
            name = str(row[2]).strip()
            activities[name] = {
                "source_key": f"{namespace}:{_number(row[0])}:ACT",
                "numero": _number(row[0]),
                "punto_procedimiento": _text(row[1]),
                "nombre": name,
                "descripcion": _text(row[3]),
                "referencia": _text(row[7]),
                "seguridad": _text(row[6]),
                "orden": order,
                "criterios": [],
            }

        criterion_order: defaultdict[str, int] = defaultdict(int)
        for row_number, row in enumerate(list(criterion_sheet.values)[4:], 5):
            if len(row) < 8 or not row[2] or not row[3]:
                continue
            activity_name = str(row[2]).strip()
            if activity_name not in activities:
                raise ValueError(
                    "Criterios individualizados:"
                    f"{row_number}: actividad no encontrada: {activity_name}"
                )
            order = criterion_order[activity_name]
            criterion_order[activity_name] += 1
            macro_keys = _macro_keys(row)
            if not macro_keys:
                raise ValueError(
                    "Criterios individualizados:"
                    f"{row_number}: criterio sin macrocompetencia"
                )
            indicators = [
                {"categoria": category, "texto": item}
                for category, column in (
                    ("tecnica", 4),
                    ("conductual", 5),
                )
                for item in _split(row[column])
            ]
            activities[activity_name]["criterios"].append(
                {
                    "source_key": (
                        f"{namespace}:{_number(row[0])}:PERF:{order + 1:02d}"
                    ),
                    "fila": row_number,
                    "descripcion": str(row[3]).strip(),
                    "referencia": _text(row[7]),
                    "orden": order,
                    "critico": False,
                    "macro_keys": sorted(macro_keys),
                    "indicadores": indicators,
                }
            )

        for activity in activities.values():
            safety = activity.pop("seguridad")
            if not safety:
                continue
            activity["criterios"].append(
                {
                    "source_key": f"{namespace}:{activity['numero']}:SAFETY",
                    "fila": 0,
                    "descripcion": safety,
                    "referencia": activity["referencia"],
                    "orden": len(activity["criterios"]),
                    "critico": True,
                    "macro_keys": ["seguridad"],
                    "indicadores": [
                        {"categoria": "seguridad", "texto": item}
                        for item in _split(safety)
                    ],
                }
            )

        criteria = [
            criterion
            for activity in activities.values()
            for criterion in activity["criterios"]
        ]
        errors = []
        if not activities:
            errors.append("No se encontraron actividades para importar")
        performance_count = sum(not item["critico"] for item in criteria)
        safety_count = sum(item["critico"] for item in criteria)
        if performance_count == 0:
            errors.append("No se encontraron criterios de desempeño")
        if safety_count != len(activities):
            errors.append(
                "Cada actividad debe tener un criterio de seguridad: "
                f"{safety_count} de {len(activities)}"
            )
        normalized = json.dumps(
            list(activities.values()), ensure_ascii=False, sort_keys=True
        ).encode()
        return {
            "activities": list(activities.values()),
            "criteria": criteria,
            "title": title,
            "source": source,
            "procedure_code": procedure,
            "namespace": namespace,
            "sheet_names": sheets,
            "errors": errors,
            "raw_source_hash": hashlib.sha256(content).hexdigest(),
            "normalized_hash": hashlib.sha256(normalized).hexdigest(),
        }
    finally:
        workbook.close()


def validate_operator_profile(
    content: bytes, db: Session, machine_id: int | None
) -> dict[str, Any]:
    try:
        profile = parse_operator_profile(content)
    except Exception as exc:
        return {
            "valido": False,
            "errores": [{"hoja": "archivo", "fila": 0, "error": str(exc)}],
            "advertencias": [],
            "resumen": {},
        }
    if machine_id is None:
        profile["errors"].append("Seleccione la máquina destino")
    elif db.query(Maquina).filter_by(id=machine_id, activo=True).first() is None:
        profile["errors"].append("La máquina seleccionada no existe o está inactiva")
    return {
        "valido": not profile["errors"],
        "errores": [
            {"hoja": "archivo", "fila": 0, "error": error}
            for error in profile["errors"]
        ],
        "advertencias": [],
        "resumen": {
            "actividades": {
                "nuevos": len(profile["activities"]),
                "actualizar": 0,
                "omitir": 0,
            },
            "criterios_desempeno": {
                "nuevos": sum(not item["critico"] for item in profile["criteria"]),
                "actualizar": 0,
                "omitir": 0,
            },
            "criterios_seguridad": {
                "nuevos": sum(item["critico"] for item in profile["criteria"]),
                "actualizar": 0,
                "omitir": 0,
            },
            "macrocompetencias": {"nuevos": 6, "actualizar": 0, "omitir": 0},
            "matrices": {"nuevos": 2, "actualizar": 0, "omitir": 0},
        },
    }


def _upsert_named(db: Session, model, name: str, **values):
    item = db.query(model).filter(model.nombre == name).first()
    if item is None:
        item = model(nombre=name, **values)
        db.add(item)
        db.flush()
    else:
        for key, value in values.items():
            setattr(item, key, value)
    return item


def _position_for_role(
    db: Session,
    machine: Maquina,
    area: Area,
    process_id: int | None,
    role: str,
    position_id: int | None = None,
) -> Puesto:
    config = ROLE_CONFIG[role]
    cargo = _upsert_named(
        db,
        Cargo,
        config["cargo"],
        descripcion=f"Cargo {config['cargo'].lower()} de máquina",
        activo=True,
    )
    position = db.get(Puesto, position_id) if position_id else None
    if position and position.tipo_puesto != role:
        raise ValueError(f"El puesto {position.nombre} no corresponde al rol {role}")
    if position and position.maquina_id not in {None, machine.id}:
        raise ValueError(f"El puesto {position.nombre} pertenece a otra máquina")
    if position is None:
        position = (
            db.query(Puesto)
            .filter_by(maquina_id=machine.id, tipo_puesto=role)
            .order_by(Puesto.id)
            .first()
        )
    if position is None:
        position = Puesto(
            codigo=generar_codigo(db, "puesto"),
            nombre=f"{config['prefix']} {machine.nombre.strip()}",
            descripcion=f"Perfil del procedimiento MP-PO-TS12-ASE-008 para {role}",
            area_id=area.id,
            cargo_id=cargo.id,
            proceso_id=process_id,
            maquina_id=machine.id,
            tipo_puesto=role,
            activo=True,
        )
        db.add(position)
        db.flush()
    else:
        position.area_id = area.id
        position.cargo_id = cargo.id
        position.proceso_id = process_id
        position.maquina_id = machine.id
        position.tipo_puesto = role
        position.activo = True

    active_machine = (
        db.query(PuestoMaquina).filter_by(puesto_id=position.id, fecha_fin=None).first()
    )
    if active_machine is None:
        db.add(
            PuestoMaquina(
                puesto_id=position.id,
                maquina_id=machine.id,
                fecha_inicio=date.today(),
            )
        )
    elif active_machine.maquina_id != machine.id:
        active_machine.fecha_fin = date.today()
        db.add(
            PuestoMaquina(
                puesto_id=position.id,
                maquina_id=machine.id,
                fecha_inicio=date.today(),
            )
        )
    return position


def _upsert_competencies(db: Session) -> dict[str, Competencia]:
    competencies: dict[str, Competencia] = {}
    for key, (name, dimension, critical) in MACRO_COMPETENCIES.items():
        competency = db.query(Competencia).filter_by(nombre=name).first()
        if competency is None:
            competency = Competencia(
                codigo=generar_codigo(db, "competencia"),
                nombre=name,
                dimension=dimension,
                critica=critical,
                nivel_sugerido=3,
                activo=True,
            )
            db.add(competency)
            db.flush()
        else:
            competency.dimension = dimension
            competency.critica = critical
            competency.nivel_sugerido = 3
            competency.activo = True
        competencies[key] = competency
    return competencies


def _create_matrix_version(
    db: Session,
    profile: dict[str, Any],
    batch: ImportacionMatriz,
    position: Puesto,
    machine: Maquina,
    area: Area,
    role: str,
    competencies: dict[str, Competencia],
    source_name: str,
    source_hash: str,
    role_settings: dict[str, Any],
    configuration: dict[str, Any],
) -> MatrizPuestoVersion:
    versions = db.query(MatrizPuestoVersion).filter_by(puesto_id=position.id)
    version_number = (
        versions.with_entities(func.max(MatrizPuestoVersion.version)).scalar() or 0
    ) + 1
    versions.filter_by(estado="publicada").update({"estado": "retirada"})
    version = MatrizPuestoVersion(
        puesto_id=position.id,
        importacion_id=batch.id,
        version=version_number,
        estado="borrador",
        source_hash=source_hash,
        importer_version=IMPORTER_VERSION,
        fuente=source_name,
    )
    db.add(version)
    db.flush()

    for activity_data in profile["activities"]:
        activity_settings = configuration.get("activities", {}).get(
            activity_data["source_key"], {}
        )
        if role not in activity_settings.get(
            "included_roles", ["operador", "ayudante"]
        ):
            continue
        activity = (
            db.query(Actividad)
            .filter_by(source_key=activity_data["source_key"])
            .first()
            or db.query(Actividad).filter_by(nombre=activity_data["nombre"]).first()
        )
        values = {
            "source_key": activity_data["source_key"],
            "nombre": activity_data["nombre"],
            "descripcion": activity_data["descripcion"],
            "punto_procedimiento": activity_data["punto_procedimiento"],
            "referencia": activity_data["referencia"],
            "orden": activity_data["orden"],
            "activo": True,
        }
        if activity is None:
            activity = Actividad(**values)
            db.add(activity)
            db.flush()
        else:
            for field, value in values.items():
                setattr(activity, field, value)
        if (
            not db.query(ActividadArea)
            .filter_by(actividad_id=activity.id, area_id=area.id)
            .first()
        ):
            db.add(ActividadArea(actividad_id=activity.id, area_id=area.id))
        if (
            not db.query(ActividadMaquina)
            .filter_by(actividad_id=activity.id, maquina_id=machine.id)
            .first()
        ):
            db.add(ActividadMaquina(actividad_id=activity.id, maquina_id=machine.id))

        assignment = PuestoActividad(
            puesto_id=position.id,
            actividad_id=activity.id,
            matriz_version_id=version.id,
        )
        db.add(assignment)
        db.flush()
        included_criteria = [
            criterion
            for criterion in activity_data["criterios"]
            if role
            in configuration.get("criteria", {})
            .get(criterion["source_key"], {})
            .get("included_roles", ["operador", "ayudante"])
        ]
        required_keys = {
            key
            for criterion in included_criteria
            for key in configuration.get("criteria", {})
            .get(criterion["source_key"], {})
            .get("macro_keys", criterion["macro_keys"])
        }
        requirements: dict[str, PuestoActividadCompetencia] = {}
        for key in required_keys:
            level = (
                role_settings["safety_level"]
                if key == "seguridad"
                else role_settings["general_level"]
            )
            requirement = PuestoActividadCompetencia(
                puesto_actividad_id=assignment.id,
                competencia_id=competencies[key].id,
                nivel_minimo=level,
            )
            db.add(requirement)
            db.flush()
            requirements[key] = requirement

        for criterion_data in included_criteria:
            criterion_settings = configuration.get("criteria", {}).get(
                criterion_data["source_key"], {}
            )
            macro_keys = criterion_settings.get(
                "macro_keys", criterion_data["macro_keys"]
            )
            criterion = (
                db.query(ActividadCriterio)
                .filter_by(source_key=criterion_data["source_key"])
                .first()
                or db.query(ActividadCriterio)
                .filter_by(
                    actividad_id=activity.id,
                    descripcion=criterion_data["descripcion"],
                )
                .first()
            )
            criterion_values = {
                "source_key": criterion_data["source_key"],
                "actividad_id": activity.id,
                "descripcion": criterion_data["descripcion"],
                "referencia": criterion_data["referencia"],
                "orden": criterion_data["orden"],
                "critico": criterion_settings.get(
                    "critical", criterion_data["critico"]
                ),
                "activo": True,
            }
            if criterion is None:
                criterion = ActividadCriterio(**criterion_values)
                db.add(criterion)
                db.flush()
            else:
                for field, value in criterion_values.items():
                    setattr(criterion, field, value)

            if role == "operador":
                desired_ids = {competencies[key].id for key in macro_keys}
                existing_ids = {link.competencia_id for link in criterion.competencias}
                for link in list(criterion.competencias):
                    if link.competencia_id not in desired_ids:
                        db.delete(link)
                for competency_id in desired_ids - existing_ids:
                    db.add(
                        ActividadCriterioCompetencia(
                            criterio_id=criterion.id,
                            competencia_id=competency_id,
                        )
                    )

            position_criterion = PuestoActividadCriterio(
                puesto_actividad_id=assignment.id,
                criterio_id=criterion.id,
                source_key=f"{role}:{criterion_data['source_key']}",
                orden=criterion_data["orden"],
                obligatorio=criterion_settings.get("required", True),
                critico=criterion_settings.get("critical", criterion_data["critico"]),
                activo=True,
            )
            db.add(position_criterion)
            db.flush()
            db.add_all(
                [
                    PuestoActividadCriterioRequisito(
                        puesto_criterio_id=position_criterion.id,
                        requisito_id=requirements[key].id,
                    )
                    for key in macro_keys
                ]
            )
            db.add_all(
                [
                    CriterioIndicador(
                        puesto_criterio_id=position_criterion.id,
                        categoria=indicator["categoria"],
                        texto=indicator["texto"],
                    )
                    for indicator in criterion_data["indicadores"]
                ]
            )
    return version


def default_profile_configuration(
    profile: dict[str, Any], machine_ids: list[int] | None = None
) -> dict[str, Any]:
    return {
        "destinations": [
            {
                "machine_id": machine_id,
                "equipment_label": None,
                "roles": {
                    "operador": {
                        "enabled": True,
                        "position_id": None,
                        "general_level": 3,
                        "safety_level": 3,
                    },
                    "ayudante": {
                        "enabled": True,
                        "position_id": None,
                        "general_level": 2,
                        "safety_level": 3,
                    },
                },
            }
            for machine_id in (machine_ids or [])
        ],
        "activities": {
            activity["source_key"]: {"included_roles": ["operador", "ayudante"]}
            for activity in profile["activities"]
        },
        "criteria": {
            criterion["source_key"]: {
                "included_roles": ["operador", "ayudante"],
                "required": True,
                "critical": criterion["critico"],
                "macro_keys": criterion["macro_keys"],
            }
            for criterion in profile["criteria"]
        },
    }


def validate_profile_configuration(
    profile: dict[str, Any], configuration: dict[str, Any], db: Session
) -> list[str]:
    errors = list(profile["errors"])
    destinations = configuration.get("destinations", [])
    machine_ids = [item.get("machine_id") for item in destinations]
    if not machine_ids:
        errors.append("Seleccione al menos una máquina destino")
    if len(machine_ids) != len(set(machine_ids)):
        errors.append("No se puede repetir una máquina destino")
    active_machine_ids = {
        item.id
        for item in db.query(Maquina)
        .filter(Maquina.id.in_(machine_ids), Maquina.activo.is_(True))
        .all()
    }
    for machine_id in machine_ids:
        if machine_id not in active_machine_ids:
            errors.append(f"La máquina {machine_id} no existe o está inactiva")
    valid_sources = {item["source_key"] for item in profile["criteria"]}
    for destination in destinations:
        roles = destination.get("roles", {})
        if not any(settings.get("enabled") for settings in roles.values()):
            errors.append(
                f"La máquina {destination.get('machine_id')} no tiene roles habilitados"
            )
        for role, settings in roles.items():
            if role not in ROLE_CONFIG or not settings.get("enabled"):
                continue
            for field in ("general_level", "safety_level"):
                level = settings.get(field)
                if not isinstance(level, int) or not 0 <= level <= 4:
                    errors.append(f"{role}.{field} debe estar entre 0 y 4")
            if settings.get("safety_level", 0) < 3:
                errors.append(f"Seguridad debe requerir al menos nivel 3 para {role}")
            position_id = settings.get("position_id")
            if position_id:
                position = db.get(Puesto, position_id)
                if position is None:
                    errors.append(f"El puesto {position_id} no existe")
                elif position.tipo_puesto != role:
                    errors.append(
                        f"El puesto {position.nombre} no corresponde al rol {role}"
                    )
                elif position.maquina_id not in {
                    None,
                    destination.get("machine_id"),
                }:
                    errors.append(
                        f"El puesto {position.nombre} pertenece a otra máquina"
                    )
    for source_key, settings in configuration.get("criteria", {}).items():
        if source_key not in valid_sources:
            errors.append(f"Criterio desconocido en configuración: {source_key}")
        if not settings.get("macro_keys"):
            errors.append(f"El criterio {source_key} no tiene macrocompetencias")
        invalid_macros = set(settings.get("macro_keys", [])) - set(MACRO_COMPETENCIES)
        if invalid_macros:
            errors.append(
                f"Macrocompetencias inválidas para {source_key}: "
                + ", ".join(sorted(invalid_macros))
            )
    for destination in destinations:
        for role, settings in destination.get("roles", {}).items():
            if not settings.get("enabled"):
                continue
            included = [
                item
                for item in profile["criteria"]
                if role
                in configuration.get("criteria", {})
                .get(item["source_key"], {})
                .get("included_roles", ["operador", "ayudante"])
            ]
            if not included:
                errors.append(f"El rol {role} no tiene criterios incluidos")
    return list(dict.fromkeys(errors))


def _configuration_hash(*values: Any) -> str:
    serialized = json.dumps(values, ensure_ascii=False, sort_keys=True).encode()
    return hashlib.sha256(serialized).hexdigest()


def _destination_hash(
    profile: dict[str, Any],
    destination: dict[str, Any],
    configuration: dict[str, Any],
) -> str:
    return _configuration_hash(
        profile["normalized_hash"],
        destination,
        configuration.get("activities"),
        configuration.get("criteria"),
        IMPORTER_VERSION,
    )


def preview_profile_configuration(
    profile: dict[str, Any], configuration: dict[str, Any], db: Session
) -> list[dict[str, Any]]:
    preview = []
    for destination in configuration.get("destinations", []):
        machine = db.get(Maquina, destination.get("machine_id"))
        if machine is None:
            continue
        destination_hash = _destination_hash(profile, destination, configuration)
        for role, settings in destination.get("roles", {}).items():
            if not settings.get("enabled"):
                continue
            position = (
                db.get(Puesto, settings.get("position_id"))
                if settings.get("position_id")
                else (
                    db.query(Puesto)
                    .filter_by(maquina_id=machine.id, tipo_puesto=role)
                    .first()
                )
            )
            role_hash = _configuration_hash(destination_hash, role, settings)
            identical = (
                db.query(MatrizPuestoVersion)
                .filter_by(
                    puesto_id=position.id,
                    source_hash=role_hash,
                    importer_version=IMPORTER_VERSION,
                )
                .first()
                if position
                else None
            )
            current = (
                db.query(MatrizPuestoVersion)
                .filter_by(puesto_id=position.id, estado="publicada")
                .first()
                if position
                else None
            )
            preview.append(
                {
                    "machine_id": machine.id,
                    "machine": machine.nombre,
                    "equipment_label": destination.get("equipment_label"),
                    "role": role,
                    "position_id": position.id if position else None,
                    "position": position.nombre if position else None,
                    "action": (
                        "sin_cambios"
                        if identical
                        else "nueva_version"
                        if current
                        else "crear_matriz"
                    ),
                    "current_version": current.version if current else None,
                }
            )
    return preview


def execute_configured_profile(
    profile: dict[str, Any],
    configuration: dict[str, Any],
    db: Session,
    source_name: str,
    user_id: int | None = None,
) -> dict[str, Any]:
    errors = validate_profile_configuration(profile, configuration, db)
    if errors:
        return {
            "valido": False,
            "errores": [
                {"hoja": "configuracion", "fila": 0, "error": error} for error in errors
            ],
            "advertencias": [],
            "resumen": {},
        }

    competencies = _upsert_competencies(db)
    created_versions: list[MatrizPuestoVersion] = []
    omitted = 0
    target_positions: set[int] = set()
    for destination in configuration["destinations"]:
        machine = db.get(Maquina, destination["machine_id"])
        active_process = (
            db.query(MaquinaProceso)
            .filter_by(maquina_id=machine.id, fecha_fin=None)
            .first()
        )
        process = active_process.proceso if active_process else None
        area = (
            process.area
            if process
            else _upsert_named(
                db,
                Area,
                AREA_NAME,
                descripcion="Área importada desde el perfil operativo",
                activo=True,
            )
        )
        destination_hash = _destination_hash(profile, destination, configuration)
        batch = (
            db.query(ImportacionMatriz)
            .filter_by(
                maquina_id=machine.id,
                normalized_hash=destination_hash,
                importer_version=IMPORTER_VERSION,
            )
            .first()
        )
        if batch is None:
            batch = ImportacionMatriz(
                maquina_id=machine.id,
                raw_source_hash=profile["raw_source_hash"],
                normalized_hash=destination_hash,
                importer_version=IMPORTER_VERSION,
                fuente=source_name,
                estado="validada",
                creado_por_usuario_id=user_id,
            )
            db.add(batch)
            db.flush()
        for role, role_settings in destination["roles"].items():
            if not role_settings.get("enabled"):
                continue
            position = _position_for_role(
                db,
                machine,
                area,
                process.id if process else None,
                role,
                role_settings.get("position_id"),
            )
            target_positions.add(position.id)
            role_hash = _configuration_hash(destination_hash, role, role_settings)
            existing = (
                db.query(MatrizPuestoVersion)
                .filter_by(
                    puesto_id=position.id,
                    source_hash=role_hash,
                    importer_version=IMPORTER_VERSION,
                )
                .first()
            )
            if existing:
                omitted += 1
                continue
            created_versions.append(
                _create_matrix_version(
                    db,
                    profile,
                    batch,
                    position,
                    machine,
                    area,
                    role,
                    competencies,
                    source_name,
                    role_hash,
                    role_settings,
                    configuration,
                )
            )
        batch.estado = "publicada"

    now = datetime.now()
    for version in created_versions:
        version.estado = "publicada"
        version.publicada_en = now
    if profile["procedure_code"] == "MP-PO-TS12-ASE-008":
        legacy = (
            db.query(Puesto)
            .filter_by(nombre=LEGACY_POSITION_NAME, maquina_id=None)
            .first()
        )
        if legacy and legacy.id not in target_positions:
            legacy.activo = False
            db.query(MatrizPuestoVersion).filter_by(
                puesto_id=legacy.id, estado="publicada"
            ).update({"estado": "retirada"})
    db.flush()
    return {
        "valido": True,
        "errores": [],
        "advertencias": [],
        "resumen": {
            "actividades": {
                "creados": len(profile["activities"]),
                "actualizados": 0,
                "omitidos": 0,
            },
            "criterios_desempeno": {
                "creados": sum(not item["critico"] for item in profile["criteria"]),
                "actualizados": 0,
                "omitidos": 0,
            },
            "criterios_seguridad": {
                "creados": sum(item["critico"] for item in profile["criteria"]),
                "actualizados": 0,
                "omitidos": 0,
            },
            "matrices": {
                "creados": len(created_versions),
                "actualizados": 0,
                "omitidos": omitted,
            },
        },
    }


def execute_operator_profile(
    content: bytes,
    db: Session,
    machine_id: int | None,
    user_id: int | None = None,
    source_name: str = "perfil_operativo.xlsx",
) -> dict[str, Any]:
    profile = parse_operator_profile(content)
    if machine_id is None:
        return validate_operator_profile(content, db, machine_id)
    configuration = default_profile_configuration(profile, [machine_id])
    result = execute_configured_profile(
        profile, configuration, db, source_name, user_id
    )
    if result["valido"]:
        db.commit()
    return result
