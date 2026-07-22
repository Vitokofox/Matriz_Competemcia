from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models import (
    Actividad,
    Area,
    AsignacionLaboral,
    Cargo,
    Competencia,
    Evaluador,
    Maquina,
    MaquinaProceso,
    Proceso,
    Puesto,
    PuestoActividad,
    PuestoActividadCompetencia,
    Supervisor,
    Trabajador,
    TrabajadorPuesto,
    Turno,
)
from app.services.codigos import generar_codigo

SHEETS: dict[str, tuple[str, ...]] = {
    "areas": ("nombre", "descripcion", "activo"),
    "cargos": ("nombre", "descripcion", "activo"),
    "turnos": ("nombre", "activo"),
    "procesos": ("nombre", "descripcion", "area", "activo"),
    "maquinas": ("nombre", "descripcion", "proceso", "fecha_inicio", "activo"),
    "actividades": ("nombre", "descripcion", "activo"),
    "competencias": ("nombre", "descripcion", "activo"),
    "puestos": (
        "nombre",
        "descripcion",
        "area",
        "proceso",
        "cargo",
        "maquina",
        "tipo_puesto",
        "activo",
    ),
    "requisitos_puesto": ("puesto", "actividad", "competencia", "nivel_minimo"),
    "supervisores": ("documento", "nombres", "apellidos", "correo", "activo"),
    "evaluadores": ("documento", "nombres", "apellidos", "correo", "activo"),
    "trabajadores": ("documento", "nombres", "apellidos", "activo"),
    "asignaciones_trabajador": (
        "trabajador_documento",
        "cargo",
        "area",
        "turno",
        "fecha_inicio",
    ),
    "trabajador_puestos": (
        "trabajador_documento",
        "puesto",
        "fecha_inicio",
    ),
}

REQUIRED: dict[str, tuple[str, ...]] = {
    "areas": ("nombre",),
    "cargos": ("nombre",),
    "turnos": ("nombre",),
    "procesos": ("nombre", "area"),
    "maquinas": ("nombre",),
    "actividades": ("nombre",),
    "competencias": ("nombre",),
    "puestos": ("nombre", "area", "cargo", "tipo_puesto"),
    "requisitos_puesto": ("puesto", "actividad", "competencia", "nivel_minimo"),
    "supervisores": ("documento", "nombres", "apellidos"),
    "evaluadores": ("documento", "nombres", "apellidos"),
    "trabajadores": ("documento", "nombres", "apellidos"),
    "asignaciones_trabajador": (
        "trabajador_documento",
        "cargo",
        "area",
        "turno",
        "fecha_inicio",
    ),
    "trabajador_puestos": ("trabajador_documento", "puesto", "fecha_inicio"),
}


def create_template() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    examples: dict[str, list[Any]] = {
        "areas": ["Producción", "Área principal de operación", True],
        "cargos": ["Operador", "Operador de máquina", True],
        "turnos": ["Día", True],
        "procesos": ["Mecanizado", "Proceso de mecanizado", "Producción", True],
        "maquinas": [
            "Torno CNC 01",
            "Torno de ejemplo",
            "Mecanizado",
            "2026-07-22",
            True,
        ],
        "actividades": ["Preparar máquina", "Preparación segura", True],
        "competencias": ["Seguridad operacional", "Trabajo seguro", True],
        "puestos": [
            "Operador de torno",
            "Puesto de ejemplo",
            "Producción",
            "Mecanizado",
            "Operador",
            "Torno CNC 01",
            "operador",
            True,
        ],
        "requisitos_puesto": [
            "Operador de torno",
            "Preparar máquina",
            "Seguridad operacional",
            4,
        ],
        "supervisores": ["11223344", "Ana", "Supervisora", "ana@example.com", True],
        "evaluadores": ["15162855", "Victor", "Valenzuela", "victor@example.com", True],
        "trabajadores": ["99887766", "Juan", "Pérez", True],
        "asignaciones_trabajador": [
            "99887766",
            "Operador",
            "Producción",
            "Día",
            "2026-07-22",
        ],
        "trabajador_puestos": ["99887766", "Operador de torno", "2026-07-22"],
    }
    for name, columns in SHEETS.items():
        sheet = workbook.create_sheet(name)
        headers = ["__ejemplo__", *columns]
        sheet.append(headers)
        sheet.append(["SI", *examples[name]])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="24704F")
        sheet.column_dimensions["A"].hidden = True
        for column in range(2, len(headers) + 1):
            sheet.column_dimensions[chr(64 + column)].width = 22
    instructions = workbook.create_sheet("LEAME", 0)
    instructions.append(["Plantilla de carga masiva", "Matriz de Competencias"])
    instructions.append(
        ["Importante", "No incluya códigos: el sistema los genera automáticamente."]
    )
    instructions.append(
        ["Ejemplos", "La fila marcada como ejemplo se ignora automáticamente."]
    )
    instructions.append(
        [
            "Modo recomendado",
            "Actualizar campos informados. Las celdas vacías no sobrescriben datos.",
        ]
    )
    instructions.append(
        [
            "Segunda opción",
            "Omitir existentes: no modifica registros encontrados por su "
            "clave natural.",
        ]
    )
    instructions.append(
        [
            "Fechas",
            "Use formato AAAA-MM-DD. Las referencias se escriben por nombre o "
            "documento.",
        ]
    )
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _bool(value: Any) -> bool:
    return str(value).strip().lower() not in {"false", "0", "no", "n", "inactivo"}


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    return date.fromisoformat(text)


def read_rows(
    content: bytes,
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    rows: dict[str, list[dict[str, Any]]] = {}
    errors: list[dict[str, Any]] = []
    for sheet_name, columns in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        values = list(sheet.values)
        if not values:
            continue
        headers = [str(value).strip().lower() if value else "" for value in values[0]]
        for index, required in enumerate(("__ejemplo__", *columns)):
            if required not in headers:
                errors.append(
                    {
                        "hoja": sheet_name,
                        "fila": 1,
                        "error": f"Falta la columna {required}",
                    }
                )
                break
        else:
            records = []
            for row_number, values_row in enumerate(values[1:], 2):
                record = {
                    headers[i]: values_row[i] if i < len(values_row) else None
                    for i in range(len(headers))
                }
                if str(record.get("__ejemplo__") or "").strip().upper() in {
                    "SI",
                    "SÍ",
                    "TRUE",
                    "EJEMPLO",
                }:
                    continue
                if not any(_text(value) for value in values_row):
                    continue
                record["__fila__"] = row_number
                records.append(record)
            rows[sheet_name] = records
    workbook.close()
    return rows, errors


def validate_import(content: bytes, db: Session) -> dict[str, Any]:
    try:
        rows, errors = read_rows(content)
    except Exception as exc:
        return {
            "valido": False,
            "errores": [
                {
                    "hoja": "archivo",
                    "fila": 0,
                    "error": f"Archivo Excel inválido: {exc}",
                }
            ],
            "resumen": {},
        }
    warnings: list[dict[str, Any]] = []
    summary = {sheet: {"nuevos": 0, "actualizar": 0, "omitir": 0} for sheet in rows}
    refs = {
        sheet: {_text(row.get("nombre")) for row in values if _text(row.get("nombre"))}
        for sheet, values in rows.items()
    }
    key_columns = {
        "areas": (Area, "nombre"),
        "cargos": (Cargo, "nombre"),
        "turnos": (Turno, "nombre"),
        "procesos": (Proceso, "nombre"),
        "maquinas": (Maquina, "nombre"),
        "actividades": (Actividad, "nombre"),
        "competencias": (Competencia, "nombre"),
        "puestos": (Puesto, "nombre"),
        "supervisores": (Supervisor, "documento"),
        "evaluadores": (Evaluador, "documento"),
        "trabajadores": (Trabajador, "documento"),
    }
    for sheet, values in rows.items():
        if sheet in key_columns:
            model, field = key_columns[sheet]
            seen: set[str] = set()
            for row in values:
                value = _text(row.get(field))
                if value in seen:
                    errors.append(
                        {
                            "hoja": sheet,
                            "fila": row["__fila__"],
                            "error": f"Registro repetido: {value}",
                        }
                    )
                seen.add(value)
                if value and not _find(db, model, field, value):
                    summary[sheet]["nuevos"] += 1
                elif value:
                    summary[sheet]["actualizar"] += 1
    for sheet, values in rows.items():
        for row in values:
            missing = [field for field in REQUIRED[sheet] if not _text(row.get(field))]
            if missing:
                errors.append(
                    {
                        "hoja": sheet,
                        "fila": row["__fila__"],
                        "error": f"Campos obligatorios vacíos: {', '.join(missing)}",
                    }
                )
                continue
            try:
                if (
                    sheet
                    in {"maquinas", "asignaciones_trabajador", "trabajador_puestos"}
                    or sheet == "puestos"
                ):
                    if sheet == "maquinas" and row.get("fecha_inicio"):
                        _date(row["fecha_inicio"])
                    if sheet == "asignaciones_trabajador":
                        _date(row["fecha_inicio"])
                    if sheet == "trabajador_puestos":
                        _date(row["fecha_inicio"])
                if (
                    sheet == "requisitos_puesto"
                    and not 1 <= int(row["nivel_minimo"]) <= 5
                ):
                    raise ValueError("nivel_minimo debe estar entre 1 y 5")
            except (ValueError, TypeError) as exc:
                errors.append(
                    {"hoja": sheet, "fila": row["__fila__"], "error": str(exc)}
                )
                continue
            if sheet == "procesos" and _text(row.get("area")) not in refs.get(
                "areas", set()
            ):
                if not _find(db, Area, "nombre", _text(row.get("area"))):
                    errors.append(
                        {
                            "hoja": sheet,
                            "fila": row["__fila__"],
                            "error": f"Área no encontrada: {row.get('area')}",
                        }
                    )
            references = {
                "maquinas": (("proceso", Proceso, "nombre"),),
                "puestos": (
                    ("area", Area, "nombre"),
                    ("proceso", Proceso, "nombre"),
                    ("cargo", Cargo, "nombre"),
                    ("maquina", Maquina, "nombre"),
                ),
                "requisitos_puesto": (
                    ("puesto", Puesto, "nombre"),
                    ("actividad", Actividad, "nombre"),
                    ("competencia", Competencia, "nombre"),
                ),
                "asignaciones_trabajador": (
                    ("trabajador_documento", Trabajador, "documento"),
                    ("cargo", Cargo, "nombre"),
                    ("area", Area, "nombre"),
                    ("turno", Turno, "nombre"),
                ),
                "trabajador_puestos": (
                    ("trabajador_documento", Trabajador, "documento"),
                    ("puesto", Puesto, "nombre"),
                ),
            }
            for field, model, lookup in references.get(sheet, ()):
                value = _text(row.get(field))
                if (
                    value
                    and not _find(db, model, lookup, value)
                    and value
                    not in refs.get(
                        next(
                            (
                                name
                                for name, pair in key_columns.items()
                                if pair == (model, lookup)
                            ),
                            "",
                        ),
                        set(),
                    )
                ):
                    errors.append(
                        {
                            "hoja": sheet,
                            "fila": row["__fila__"],
                            "error": f"Referencia no encontrada: {field}={value}",
                        }
                    )
    return {
        "valido": not errors,
        "errores": errors,
        "advertencias": warnings,
        "resumen": summary,
    }


def _find(db: Session, model: Any, field: str, value: Any):
    return db.query(model).filter(getattr(model, field) == value).first()


def _upsert(
    db: Session,
    model: Any,
    key: str,
    row: dict[str, Any],
    fields: tuple[str, ...],
    entity: str | None,
    skip: bool,
):
    value = _text(row.get(key))
    item = _find(db, model, key, value)
    if item and skip:
        return item, False, False
    created = False
    if not item:
        data = {
            field: _text(row.get(field))
            for field in fields
            if _text(row.get(field)) is not None
        }
        if "activo" in fields and row.get("activo") is not None:
            data["activo"] = _bool(row["activo"])
        if entity:
            data["codigo"] = generar_codigo(db, entity)
        item = model(**data)
        db.add(item)
        db.flush()
        created = True
    else:
        for field in fields:
            value = _text(row.get(field))
            if value is not None:
                setattr(item, field, _bool(value) if field == "activo" else value)
    return item, True, created


def execute_import(content: bytes, db: Session, skip_existing: bool) -> dict[str, Any]:
    validation = validate_import(content, db)
    if not validation["valido"]:
        return validation
    rows, _ = read_rows(content)
    counts = {sheet: {"creados": 0, "actualizados": 0, "omitidos": 0} for sheet in rows}

    def process(
        sheet: str,
        model: Any,
        key: str,
        fields: tuple[str, ...],
        entity: str | None = None,
    ):
        for row in rows.get(sheet, []):
            _, changed, created = _upsert(
                db, model, key, row, fields, entity, skip_existing
            )
            if not changed:
                counts[sheet]["omitidos"] += 1
            elif created:
                counts[sheet]["creados"] += 1
            else:
                counts[sheet]["actualizados"] += 1

    process("areas", Area, "nombre", ("nombre", "descripcion", "activo"))
    process("cargos", Cargo, "nombre", ("nombre", "descripcion", "activo"))
    process("turnos", Turno, "nombre", ("nombre", "activo"))
    process("actividades", Actividad, "nombre", ("nombre", "descripcion", "activo"))
    process(
        "competencias",
        Competencia,
        "nombre",
        ("nombre", "descripcion", "activo"),
        "competencia",
    )
    process(
        "supervisores",
        Supervisor,
        "documento",
        ("documento", "nombres", "apellidos", "correo", "activo"),
        "supervisor",
    )
    process(
        "evaluadores",
        Evaluador,
        "documento",
        ("documento", "nombres", "apellidos", "correo", "activo"),
        "evaluador",
    )
    process(
        "trabajadores",
        Trabajador,
        "documento",
        ("documento", "nombres", "apellidos", "activo"),
        "trabajador",
    )
    for row in rows.get("procesos", []):
        area = _find(db, Area, "nombre", _text(row["area"]))
        item, changed, _ = _upsert(
            db,
            Proceso,
            "nombre",
            row,
            ("nombre", "descripcion", "activo"),
            "proceso",
            skip_existing,
        )
        if not changed:
            continue
        item.area_id = area.id
        db.flush()
    for row in rows.get("maquinas", []):
        item, changed, _ = _upsert(
            db,
            Maquina,
            "nombre",
            row,
            ("nombre", "descripcion", "activo"),
            "maquina",
            skip_existing,
        )
        if not changed:
            continue
        process = _find(db, Proceso, "nombre", _text(row.get("proceso")))
        if process and not _find(db, MaquinaProceso, "maquina_id", item.id):
            db.add(
                MaquinaProceso(
                    maquina_id=item.id,
                    proceso_id=process.id,
                    fecha_inicio=_date(row.get("fecha_inicio")) or date.today(),
                )
            )
    for row in rows.get("puestos", []):
        data = {
            "nombre": _text(row["nombre"]),
            "descripcion": _text(row.get("descripcion")),
            "tipo_puesto": _text(row["tipo_puesto"]),
            "activo": _bool(row.get("activo", True)),
        }
        item = _find(db, Puesto, "nombre", data["nombre"])
        if not item:
            item = Puesto(codigo=generar_codigo(db, "puesto"), **data)
            db.add(item)
            db.flush()
        elif skip_existing:
            continue
        elif not skip_existing:
            for key, value in data.items():
                if value is not None:
                    setattr(item, key, value)
        item.area_id = _find(db, Area, "nombre", _text(row["area"])).id
        item.cargo_id = _find(db, Cargo, "nombre", _text(row["cargo"])).id
        process = _find(db, Proceso, "nombre", _text(row.get("proceso")))
        item.proceso_id = process.id if process else None
        machine = _find(db, Maquina, "nombre", _text(row.get("maquina")))
        item.maquina_id = machine.id if machine else None
    for row in rows.get("requisitos_puesto", []):
        position = _find(db, Puesto, "nombre", _text(row["puesto"]))
        activity = _find(db, Actividad, "nombre", _text(row["actividad"]))
        competency = _find(db, Competencia, "nombre", _text(row["competencia"]))
        assignment = (
            db.query(PuestoActividad)
            .filter_by(puesto_id=position.id, actividad_id=activity.id)
            .first()
        )
        if not assignment:
            assignment = PuestoActividad(
                puesto_id=position.id, actividad_id=activity.id
            )
            db.add(assignment)
            db.flush()
        requirement = (
            db.query(PuestoActividadCompetencia)
            .filter_by(puesto_actividad_id=assignment.id, competencia_id=competency.id)
            .first()
        )
        if not requirement:
            db.add(
                PuestoActividadCompetencia(
                    puesto_actividad_id=assignment.id,
                    competencia_id=competency.id,
                    nivel_minimo=int(row["nivel_minimo"]),
                )
            )
        elif not skip_existing and row.get("nivel_minimo") is not None:
            requirement.nivel_minimo = int(row["nivel_minimo"])
    for row in rows.get("asignaciones_trabajador", []):
        worker = _find(db, Trabajador, "documento", _text(row["trabajador_documento"]))
        assignment = (
            db.query(AsignacionLaboral)
            .filter_by(trabajador_id=worker.id, fecha_fin=None)
            .first()
        )
        if assignment and skip_existing:
            continue
        if assignment:
            assignment.fecha_fin = _date(row["fecha_inicio"])
        db.add(
            AsignacionLaboral(
                trabajador_id=worker.id,
                cargo_id=_find(db, Cargo, "nombre", _text(row["cargo"])).id,
                area_id=_find(db, Area, "nombre", _text(row["area"])).id,
                turno_id=_find(db, Turno, "nombre", _text(row["turno"])).id,
                fecha_inicio=_date(row["fecha_inicio"]),
            )
        )
    for row in rows.get("trabajador_puestos", []):
        worker = _find(db, Trabajador, "documento", _text(row["trabajador_documento"]))
        position = _find(db, Puesto, "nombre", _text(row["puesto"]))
        active = (
            db.query(TrabajadorPuesto)
            .filter_by(trabajador_id=worker.id, puesto_id=position.id, fecha_fin=None)
            .first()
        )
        if not active:
            db.add(
                TrabajadorPuesto(
                    trabajador_id=worker.id,
                    puesto_id=position.id,
                    fecha_inicio=_date(row["fecha_inicio"]),
                )
            )
    db.commit()
    return {"valido": True, "errores": [], "advertencias": [], "resumen": counts}
