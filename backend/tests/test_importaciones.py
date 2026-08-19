from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401
from app.api.routes import load_evaluation_criteria
from app.api.schemas import EvaluacionCriterioCreate
from app.db.session import Base
from app.models import (
    Actividad,
    ActividadCriterio,
    Area,
    Evaluacion,
    ImportacionMatriz,
    Maquina,
    MatrizPuestoVersion,
    PuestoActividad,
    PuestoActividadCompetencia,
    PuestoActividadCriterio,
)
from app.services.importacion_perfil_operador import (
    default_profile_configuration,
    execute_configured_profile,
    parse_operator_profile,
)
from app.services.importaciones import create_template, execute_import, validate_import


def workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "areas"
    sheet.append(["__ejemplo__", "nombre", "descripcion", "activo"])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def unknown_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Hoja desconocida"
    workbook.active.append(["dato"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.fixture
def db() -> Session:
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


def test_template_has_instructions_and_examples() -> None:
    workbook = load_workbook(BytesIO(create_template()), read_only=True)
    assert workbook.sheetnames[0] == "LEAME"
    assert "areas" in workbook.sheetnames
    assert list(workbook["areas"].values)[1][0] == "SI"


def test_unknown_workbook_is_rejected(db) -> None:
    result = validate_import(unknown_workbook_bytes(), db)
    assert not result["valido"]
    assert "hojas reconocidas" in result["errores"][0]["error"]


def test_import_updates_informed_fields_and_can_skip(db) -> None:
    first = workbook_bytes([[None, "Producción", "Inicial", True]])
    assert validate_import(first, db)["valido"]
    execute_import(first, db, skip_existing=False)
    second = workbook_bytes([[None, "Producción", "Actualizada", None]])
    execute_import(second, db, skip_existing=False)
    area = db.query(Area).one()
    assert (area.nombre, area.descripcion, area.activo) == (
        "Producción",
        "Actualizada",
        True,
    )
    execute_import(
        workbook_bytes([[None, "Producción", "No cambia", True]]),
        db,
        skip_existing=True,
    )
    assert (
        db.execute(text("SELECT descripcion FROM areas")).scalar_one() == "Actualizada"
    )


def test_perfil_operador_crea_checklist_versionado_y_calcula_minimo(db) -> None:
    path = (
        Path(__file__).parents[2]
        / "Actividades"
        / "Analisis_Puesto_Operador_Linea_Principal_criterios_individualizados.xlsx"
    )
    content = path.read_bytes()
    profile = parse_operator_profile(content)
    assert (len(profile["activities"]), len(profile["criteria"])) == (23, 100)
    assert profile["errors"] == []

    machine = Maquina(codigo="MAQ-TEST", nombre="Máquina principal")
    db.add(machine)
    db.flush()
    assert not validate_import(content, db)["valido"]
    assert validate_import(content, db, machine.id)["valido"]

    result = execute_import(content, db, skip_existing=False, machine_id=machine.id)
    assert result["valido"]
    versions = db.query(MatrizPuestoVersion).all()
    assert len(versions) == 2
    assert {item.puesto.tipo_puesto for item in versions} == {"operador", "ayudante"}
    assert all(item.estado == "publicada" for item in versions)
    assert all(item.puesto.maquina_id == machine.id for item in versions)
    assert all(len(item.actividades) == 23 for item in versions)
    criteria = db.query(PuestoActividadCriterio).all()
    assert len(criteria) == 200
    assert all(item.requisitos for item in criteria)

    helper = next(item for item in versions if item.puesto.tipo_puesto == "ayudante")
    helper_requirements = (
        db.query(PuestoActividadCompetencia)
        .join(PuestoActividad)
        .filter(PuestoActividad.matriz_version_id == helper.id)
        .all()
    )
    assert {
        item.nivel_minimo for item in helper_requirements if item.competencia.critica
    } == {3}
    assert 2 in {
        item.nivel_minimo
        for item in helper_requirements
        if not item.competencia.critica
    }

    version = next(item for item in versions if item.puesto.tipo_puesto == "operador")
    criteria = (
        db.query(PuestoActividadCriterio)
        .join(PuestoActividad)
        .filter(PuestoActividad.matriz_version_id == version.id)
        .all()
    )
    critical = next(item for item in criteria if item.critico)
    evaluation = Evaluacion(
        trabajador_id=1,
        puesto_id=version.puesto_id,
        supervisor_id=1,
        fecha=date.today(),
        matriz_version_id=version.id,
    )
    responses = [
        EvaluacionCriterioCreate(
            puesto_criterio_id=item.id,
            nivel_obtenido=2 if item.id == critical.id else 4,
            evidencia="Requiere apoyo" if item.id == critical.id else None,
        )
        for item in criteria
    ]
    load_evaluation_criteria(db, evaluation, responses)
    affected = {link.requisito_id for link in critical.requisitos}
    assert all(
        detail.nivel_obtenido == 2
        for detail in evaluation.detalles
        if detail.requisito_id in affected
    )
    assert any(item.critico_incumplido for item in evaluation.criterios)

    repeated = execute_import(content, db, skip_existing=False, machine_id=machine.id)
    assert repeated["resumen"]["matrices"]["omitidos"] == 2

    second_machine = Maquina(codigo="MAQ-TEST-2", nombre="Segunda máquina")
    db.add(second_machine)
    db.flush()
    second_result = execute_import(
        content, db, skip_existing=False, machine_id=second_machine.id
    )
    assert second_result["resumen"]["matrices"]["creados"] == 2
    assert db.query(MatrizPuestoVersion).count() == 4
    assert db.query(ImportacionMatriz).count() == 2
    assert db.query(Actividad).count() == 23
    assert db.query(ActividadCriterio).count() == 100
    assert all(item.source_key for item in db.query(Actividad).all())
    assert all(item.source_key for item in db.query(ActividadCriterio).all())


def test_perfil_canteadoras_configura_dos_maquinas_y_cuatro_matrices(db) -> None:
    path = (
        Path(__file__).parents[2]
        / "Actividades"
        / "Analisis_Cargo_Operador_Canteadoras_criterios_individualizados.xlsx"
    )
    profile = parse_operator_profile(path.read_bytes())
    assert profile["procedure_code"] == "MP-PO-TS12-ASE-009"
    assert len(profile["activities"]) == 10
    assert len(profile["criteria"]) == 54

    edgar = Maquina(codigo="CT1", nombre="Canteadora Edgar")
    exceltec = Maquina(codigo="CT2", nombre="Canteadora Exceltec")
    db.add_all([edgar, exceltec])
    db.flush()
    configuration = default_profile_configuration(profile, [edgar.id, exceltec.id])
    configuration["destinations"][0]["equipment_label"] = "CT1"
    configuration["destinations"][1]["equipment_label"] = "CT2"

    result = execute_configured_profile(profile, configuration, db, path.name)
    assert result["valido"]
    versions = db.query(MatrizPuestoVersion).all()
    assert len(versions) == 4
    assert {item.puesto.maquina_id for item in versions} == {edgar.id, exceltec.id}
    assert all(len(item.actividades) == 10 for item in versions)
    assert all(
        sum(len(activity.criterios_evaluables) for activity in item.actividades) == 54
        for item in versions
    )

    repeated = execute_configured_profile(profile, configuration, db, path.name)
    assert repeated["resumen"]["matrices"]["omitidos"] == 4
